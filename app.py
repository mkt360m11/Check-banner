from __future__ import annotations

import csv
import io
import os
import re
import time
from datetime import datetime, timezone
from threading import Lock
from urllib.parse import urlparse

import requests
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, render_template, request, send_file

from banner_checker import run_banner_check
from history_store import HistoryStore

load_dotenv()

PORT = int(os.getenv("BANNER_QC_PORT", "8011"))
HISTORY_DIR = os.getenv("BANNER_QC_HISTORY_DIR", "history")
DOMAIN_API = os.getenv("BANNER_DOMAIN_API", "").strip()
DOMAIN_API_TIMEOUT = max(3.0, min(float(os.getenv("BANNER_DOMAIN_API_TIMEOUT", "12")), 60.0))
DOMAIN_API_RETRIES = max(1, min(int(os.getenv("BANNER_DOMAIN_API_RETRIES", "2")), 5))
DOMAIN_API_RETRY_DELAY = max(0.0, min(float(os.getenv("BANNER_DOMAIN_API_RETRY_DELAY", "0.8")), 10.0))

app = Flask(__name__, template_folder="templates")
store = HistoryStore(HISTORY_DIR)
domain_source_lock = Lock()
domain_source_cache: set[str] = set()
domain_source_state: dict[str, str | int | float] = {
    "status": "idle",
    "last_sync_at": "",
    "last_error": "",
    "source": "cache",
    "total": 0,
}


def normalize_domain(url_or_domain: str) -> str:
    value = (url_or_domain or "").strip().lower()
    if not value:
        return ""
    if value.startswith("http://") or value.startswith("https://"):
        return urlparse(value).netloc.replace("www.", "")
    return value.replace("www.", "")


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def extract_domains_from_payload(payload: object) -> set[str]:
    items: list[object] = []

    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        for key in ("domains", "data", "items", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                items = value
                break
        if not items and any(key in payload for key in ("domain", "url", "host", "site", "main_domain")):
            items = [payload]

    normalized: set[str] = set()
    for item in items:
        if isinstance(item, str):
            domain = normalize_domain(item)
            if domain:
                normalized.add(domain)
            continue

        if isinstance(item, dict):
            candidate = (
                item.get("domain")
                or item.get("url")
                or item.get("host")
                or item.get("site")
                or item.get("main_domain")
                or item.get("destination_domain")
            )
            domain = normalize_domain(str(candidate or ""))
            if domain:
                normalized.add(domain)

    return normalized


def action_call_domain_api() -> tuple[set[str], str]:
    if not DOMAIN_API:
        return set(), "domain_api_not_configured"

    last_error = ""
    for attempt in range(1, DOMAIN_API_RETRIES + 1):
        try:
            response = requests.get(DOMAIN_API, timeout=DOMAIN_API_TIMEOUT)
            response.raise_for_status()
            payload = response.json()
            domains = extract_domains_from_payload(payload)
            return domains, "ok"
        except Exception as exc:
            last_error = str(exc)
            if attempt < DOMAIN_API_RETRIES and DOMAIN_API_RETRY_DELAY > 0:
                time.sleep(DOMAIN_API_RETRY_DELAY)

    return set(), last_error or "domain_api_unknown_error"


def fetch_latest_domains(force: bool = False) -> set[str]:
    with domain_source_lock:
        if domain_source_cache and not force:
            return set(domain_source_cache)

        domains, call_status = action_call_domain_api()

        if call_status == "ok":
            domain_source_cache.clear()
            domain_source_cache.update(domains)
            domain_source_state["status"] = "ok"
            domain_source_state["source"] = "api_live"
            domain_source_state["last_error"] = ""
            domain_source_state["last_sync_at"] = utc_now_iso()
            domain_source_state["total"] = len(domain_source_cache)
            return set(domain_source_cache)

        if not DOMAIN_API:
            domain_source_state["status"] = "no_api"
            domain_source_state["source"] = "fallback_empty"
            domain_source_state["last_error"] = "BANNER_DOMAIN_API is empty"
            domain_source_state["last_sync_at"] = utc_now_iso()
            domain_source_state["total"] = len(domain_source_cache)
            return set(domain_source_cache)

        domain_source_state["status"] = "api_error"
        domain_source_state["source"] = "cache_fallback" if domain_source_cache else "fallback_empty"
        domain_source_state["last_error"] = call_status
        domain_source_state["last_sync_at"] = utc_now_iso()
        domain_source_state["total"] = len(domain_source_cache)
        return set(domain_source_cache)


@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/domain_source")
def domain_source():
    latest_domains = sorted(fetch_latest_domains(force=False))
    return jsonify(
        {
            "domain_api": DOMAIN_API,
            "timeout_seconds": DOMAIN_API_TIMEOUT,
            "retries": DOMAIN_API_RETRIES,
            "retry_delay_seconds": DOMAIN_API_RETRY_DELAY,
            "total": len(latest_domains),
            "domains": latest_domains[:50],
            "state": domain_source_state,
        }
    )


@app.route("/api/domain_source/refresh", methods=["POST"])
def refresh_domain_source():
    latest_domains = fetch_latest_domains(force=True)
    return jsonify(
        {
            "status": "ok",
            "total": len(latest_domains),
            "state": domain_source_state,
        }
    )


@app.route("/api/domain_source/action", methods=["POST"])
def action_domain_source():
    latest_domains = fetch_latest_domains(force=True)
    return jsonify(
        {
            "status": "ok",
            "action": "call_domain_api",
            "total": len(latest_domains),
            "state": domain_source_state,
        }
    )


PROXY_FILE = os.path.join(os.path.dirname(__file__), "proxies.txt")


def _load_proxy_file() -> list[str]:
    try:
        with open(PROXY_FILE, "r") as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        return []


def _save_proxy_file(proxies: list[str]) -> None:
    with open(PROXY_FILE, "w") as f:
        f.write("\n".join(proxies))


@app.route("/api/proxies", methods=["GET"])
def get_proxies():
    return jsonify({"proxies": _load_proxy_file()})


@app.route("/api/proxies", methods=["POST"])
def save_proxies():
    data = request.get_json(silent=True) or {}
    raw = data.get("proxies", "")
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    _save_proxy_file(lines)
    return jsonify({"saved": len(lines)})


@app.route("/api/banner/check", methods=["POST"])
def banner_check():
    payload = request.get_json(silent=True) or {}

    # Merge saved proxy list if UI textarea is empty
    if not str(payload.get("proxy", "")).strip():
        saved = _load_proxy_file()
        if saved:
            payload["proxy"] = "\n".join(saved)

    latest_domains = fetch_latest_domains(force=False)
    result = run_banner_check(payload, latest_domains=latest_domains)

    # Remove dead proxies from persistent file
    health = result.get("proxy_health") or []
    if health:
        alive = [r["raw"] for r in health if r.get("alive")]
        current = _load_proxy_file()
        dead = {r["raw"] for r in health if not r.get("alive")}
        updated = [p for p in current if p not in dead]
        # Add newly confirmed alive proxies not yet in file
        for p in alive:
            if p not in updated:
                updated.append(p)
        _save_proxy_file(updated)
        result["proxy_list_updated"] = {"removed": list(dead), "remaining": len(updated)}

    filename = store.save(result)
    result["history_file"] = filename
    return jsonify(result)


@app.route("/api/history")
def history_list():
    return jsonify({"history": store.list_files(limit=100)})


@app.route("/api/history/<path:filename>")
def history_detail(filename: str):
    try:
        data = store.load(filename)
        return jsonify(data)
    except FileNotFoundError:
        return jsonify({"error": "history_not_found"}), 404


_EXPORT_COLUMNS = [
    "run_at", "site", "refresh_no", "banner_index", "position",
    "width", "height", "selector", "domain",
    "link", "destination_link", "redirect_target", "final_destination_link",
    "http_status", "redirect_http_status", "redirect_resolved_by",
    "qc_status", "qc_reasons",
]

_SITE_COLUMNS = [
    "run_at", "site", "banner_total", "pass_banners", "fail_cases", "site_status",
]


def _parse_run_at(filename: str) -> str:
    m = re.search(r"(\d{8})_(\d{6})", filename)
    if m:
        d, t = m.group(1), m.group(2)
        return f"{d[:4]}-{d[4:6]}-{d[6:8]} {t[:2]}:{t[2:4]}:{t[4:6]}"
    return filename


def _flatten_banners(data: dict, filename: str) -> list[dict]:
    run_at = _parse_run_at(filename)
    rows = []
    for site_result in data.get("results", []):
        site = site_result.get("site", "")
        for refresh in site_result.get("refreshes", []):
            for banner in refresh.get("banners", []):
                rows.append({
                    "run_at": run_at,
                    "site": site,
                    "refresh_no": refresh.get("refresh", ""),
                    "banner_index": banner.get("index", ""),
                    "position": banner.get("position", ""),
                    "width": banner.get("w", ""),
                    "height": banner.get("h", ""),
                    "selector": banner.get("source_selector", ""),
                    "domain": banner.get("domain", ""),
                    "link": banner.get("link", ""),
                    "destination_link": banner.get("destination_link", ""),
                    "redirect_target": banner.get("redirect_target", ""),
                    "final_destination_link": banner.get("final_destination_link", ""),
                    "http_status": banner.get("http_status", ""),
                    "redirect_http_status": banner.get("redirect_http_status", ""),
                    "redirect_resolved_by": banner.get("redirect_resolved_by", ""),
                    "qc_status": banner.get("qc_status", ""),
                    "qc_reasons": "|".join(banner.get("qc_reasons") or []),
                })
    return rows


def _flatten_sites(data: dict, filename: str) -> list[dict]:
    run_at = _parse_run_at(filename)
    rows = []
    for site_result in data.get("results", []):
        rows.append({
            "run_at": run_at,
            "site": site_result.get("site", ""),
            "banner_total": site_result.get("banner_total", 0),
            "pass_banners": site_result.get("pass_banners", 0),
            "fail_cases": site_result.get("fail_cases", 0),
            "site_status": site_result.get("site_status", ""),
        })
    return rows


@app.route("/api/export/csv/<path:filename>")
def export_csv(filename: str):
    try:
        data = store.load(filename)
    except FileNotFoundError:
        return jsonify({"error": "history_not_found"}), 404

    rows = _flatten_banners(data, filename)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=_EXPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

    download_name = filename.replace(".json", "_banner_detail.csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


@app.route("/api/export/excel/<path:filename>")
def export_excel(filename: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl_not_installed"}), 500

    try:
        data = store.load(filename)
    except FileNotFoundError:
        return jsonify({"error": "history_not_found"}), 404

    banner_rows = _flatten_banners(data, filename)
    site_rows = _flatten_sites(data, filename)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill("solid", fgColor="D6F5D6")
    fail_fill = PatternFill("solid", fgColor="FFD6D6")

    def _write_sheet(ws, columns, rows, status_col):
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([row.get(col, "") for col in columns])
            fill = pass_fill if row.get(status_col) == "PASS" else fail_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    ws_banner = wb.active
    ws_banner.title = "Banner Detail"
    _write_sheet(ws_banner, _EXPORT_COLUMNS, banner_rows, "qc_status")

    ws_site = wb.create_sheet("Site Summary")
    _write_sheet(ws_site, _SITE_COLUMNS, site_rows, "site_status")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    download_name = filename.replace(".json", "_report.xlsx")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=False)
